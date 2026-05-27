"""
Cyber Incident Alert Engine
============================
Lit CyberDashboard_Data.xlsx, compare les dernières valeurs aux seuils
définis dans l'onglet Configuration, et envoie un mail HTML structuré
si un seuil est dépassé.

Usage:
    python alert_engine.py                        # analyse + envoi mail si alerte
    python alert_engine.py --dry-run              # analyse sans envoi mail
    python alert_engine.py --file autre.xlsx      # fichier personnalisé

Dépendances:
    pip install openpyxl
"""

import argparse
import datetime
import smtplib
import sys
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import openpyxl

# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────
DEFAULT_FILE = "CyberDashboard_Data.xlsx"

# Mapping: (feuille Données col index 1-based) → (ligne seuil dans Configuration)
INDICATORS = [
    {"name": "Mails reçus",              "data_col": 2,  "cfg_row": 5},
    {"name": "Requêtes réseau",          "data_col": 3,  "cfg_row": 6},
    {"name": "Tentatives de connexion",  "data_col": 4,  "cfg_row": 7},
    {"name": "Alertes IDS",              "data_col": 6,  "cfg_row": 8},
    {"name": "Trafic entrant (MB)",      "data_col": 7,  "cfg_row": 9},
    {"name": "Connexions refusées",      "data_col": 9,  "cfg_row": 10},
    {"name": "Processus suspects",       "data_col": 10, "cfg_row": 11},
]

# ─────────────────────────────────────────────────────────────────────────────
# Excel reading
# ─────────────────────────────────────────────────────────────────────────────

def load_workbook_data(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    return wb


def read_config(wb) -> dict:
    ws = wb["Configuration"]
    cfg = {
        "smtp_server":   str(ws["B14"].value or "smtp.gmail.com"),
        "smtp_port":     int(ws["B15"].value or 587),
        "sender":        str(ws["B16"].value or ""),
        "password":      str(ws["B17"].value or ""),
        "recipient":     str(ws["B18"].value or ""),
        "company":       str(ws["B19"].value or "MonEntreprise"),
    }
    return cfg


def read_thresholds(wb, indicators: list) -> list:
    ws = wb["Configuration"]
    enriched = []
    for ind in indicators:
        row = ind["cfg_row"]
        enriched.append({
            **ind,
            "warn":  float(ws.cell(row, 2).value or 0),
            "crit":  float(ws.cell(row, 3).value or 0),
        })
    return enriched


def read_last_two_values(wb, indicators: list) -> list:
    ws = wb["Données"]
    max_row = ws.max_row
    results = []
    for ind in indicators:
        col = ind["data_col"]
        val_now  = ws.cell(max_row, col).value or 0
        val_prev = ws.cell(max_row - 1, col).value or 0
        timestamp = ws.cell(max_row, 1).value
        results.append({
            **ind,
            "value":     float(val_now),
            "prev":      float(val_prev),
            "timestamp": timestamp,
            "variation": ((float(val_now) - float(val_prev)) / float(val_prev) * 100)
                         if val_prev else 0,
        })
    return results

# ─────────────────────────────────────────────────────────────────────────────
# Anomaly detection
# ─────────────────────────────────────────────────────────────────────────────

def classify(value, warn, crit) -> str:
    if value >= crit:
        return "CRITIQUE"
    if value >= warn:
        return "AVERTISSEMENT"
    return "OK"


def detect_anomalies(data: list) -> list:
    alerts = []
    for item in data:
        level = classify(item["value"], item["warn"], item["crit"])
        if level != "OK":
            alerts.append({**item, "level": level})
    return alerts

# ─────────────────────────────────────────────────────────────────────────────
# Email builder
# ─────────────────────────────────────────────────────────────────────────────

LEVEL_COLOR = {
    "CRITIQUE":       ("#da3633", "#ffeef0"),
    "AVERTISSEMENT":  ("#d29922", "#fffbdd"),
}

def build_html_email(alerts: list, company: str, all_data: list) -> str:
    now = datetime.datetime.now().strftime("%d/%m/%Y à %H:%M:%S")
    has_critical = any(a["level"] == "CRITIQUE" for a in alerts)
    banner_color = "#da3633" if has_critical else "#d29922"
    banner_label = "🔴 ALERTE CRITIQUE" if has_critical else "🟡 AVERTISSEMENT"

    rows_alert = ""
    for a in alerts:
        color, bg = LEVEL_COLOR[a["level"]]
        var_sign = "+" if a["variation"] >= 0 else ""
        rows_alert += f"""
        <tr style="background:{bg};">
          <td style="padding:10px 14px;border-bottom:1px solid #e1e4e8;font-weight:600;color:#0d1117;">{a['name']}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #e1e4e8;text-align:center;font-weight:700;color:{color};font-size:16px;">{a['value']:,.0f}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #e1e4e8;text-align:center;color:#586069;">{a['warn']:,.0f}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #e1e4e8;text-align:center;color:{color};font-weight:600;">{a['crit']:,.0f}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #e1e4e8;text-align:center;">
            <span style="background:{color};color:white;padding:3px 10px;border-radius:12px;font-size:12px;font-weight:700;">{a['level']}</span>
          </td>
          <td style="padding:10px 14px;border-bottom:1px solid #e1e4e8;text-align:center;color:{'#da3633' if a['variation']>0 else '#22863a'};font-weight:600;">
            {var_sign}{a['variation']:.1f}%
          </td>
        </tr>"""

    # Summary table (all indicators)
    rows_all = ""
    for item in all_data:
        level = classify(item["value"], item["warn"], item["crit"])
        dot = {"CRITIQUE": "🔴", "AVERTISSEMENT": "🟡", "OK": "🟢"}[level]
        rows_all += f"""
        <tr>
          <td style="padding:7px 12px;border-bottom:1px solid #eaecef;font-size:13px;">{item['name']}</td>
          <td style="padding:7px 12px;border-bottom:1px solid #eaecef;text-align:center;font-size:13px;font-weight:600;">{item['value']:,.0f}</td>
          <td style="padding:7px 12px;border-bottom:1px solid #eaecef;text-align:center;font-size:13px;">{dot} {level}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="fr">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f6f8fa;font-family:Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f6f8fa;padding:24px 0;">
<tr><td align="center">
<table width="640" cellpadding="0" cellspacing="0" style="background:#ffffff;border:1px solid #e1e4e8;border-radius:8px;overflow:hidden;">

  <!-- Header -->
  <tr><td style="background:{banner_color};padding:20px 28px;">
    <p style="margin:0;font-size:22px;font-weight:700;color:#ffffff;">{banner_label}</p>
    <p style="margin:6px 0 0;font-size:13px;color:rgba(255,255,255,.85);">
      Système de surveillance — {company} &nbsp;|&nbsp; {now}
    </p>
  </td></tr>

  <!-- Alert detail -->
  <tr><td style="padding:24px 28px 8px;">
    <p style="margin:0 0 14px;font-size:15px;font-weight:600;color:#0d1117;">
      {len(alerts)} indicateur(s) dépassant les seuils configurés :
    </p>
    <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #e1e4e8;border-radius:6px;overflow:hidden;">
      <thead>
        <tr style="background:#f6f8fa;">
          <th style="padding:10px 14px;text-align:left;font-size:12px;color:#586069;border-bottom:1px solid #e1e4e8;">INDICATEUR</th>
          <th style="padding:10px 14px;font-size:12px;color:#586069;border-bottom:1px solid #e1e4e8;">VALEUR</th>
          <th style="padding:10px 14px;font-size:12px;color:#586069;border-bottom:1px solid #e1e4e8;">SEUIL AVERT.</th>
          <th style="padding:10px 14px;font-size:12px;color:#586069;border-bottom:1px solid #e1e4e8;">SEUIL CRIT.</th>
          <th style="padding:10px 14px;font-size:12px;color:#586069;border-bottom:1px solid #e1e4e8;">SÉVÉRITÉ</th>
          <th style="padding:10px 14px;font-size:12px;color:#586069;border-bottom:1px solid #e1e4e8;">VARIATION</th>
        </tr>
      </thead>
      <tbody>{rows_alert}</tbody>
    </table>
  </td></tr>

  <!-- Interpretation -->
  <tr><td style="padding:16px 28px;">
    <div style="background:#f1f8ff;border-left:4px solid #0366d6;padding:12px 16px;border-radius:0 4px 4px 0;">
      <p style="margin:0;font-size:13px;font-weight:600;color:#0366d6;">Interprétation automatique :</p>
      <ul style="margin:8px 0 0;padding-left:18px;font-size:13px;color:#0d1117;line-height:1.7;">
        {"<li>Hausse des <b>requêtes réseau</b> → possible attaque <b>DDoS</b>. Vérifier les logs firewall.</li>" if any(a['name']=='Requêtes réseau' for a in alerts) else ""}
        {"<li>Hausse des <b>mails reçus</b> → campagne de <b>phishing</b> probable. Isoler et analyser les messages.</li>" if any(a['name']=='Mails reçus' for a in alerts) else ""}
        {"<li>Hausse des <b>tentatives de connexion</b> → attaque <b>brute-force</b> en cours. Bloquer les IP sources.</li>" if any(a['name']=='Tentatives de connexion' for a in alerts) else ""}
        {"<li>Hausse des <b>alertes IDS</b> → intrusion potentiellement <b>active</b>. Contacter l'équipe SOC.</li>" if any(a['name']=='Alertes IDS' for a in alerts) else ""}
        {"<li>Hausse des <b>processus suspects</b> → possible <b>malware</b> en exécution. Lancer analyse forensique.</li>" if any(a['name']=='Processus suspects' for a in alerts) else ""}
      </ul>
    </div>
  </td></tr>

  <!-- All indicators summary -->
  <tr><td style="padding:4px 28px 24px;">
    <p style="font-size:13px;font-weight:600;color:#0d1117;margin:0 0 8px;">Tous les indicateurs :</p>
    <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #e1e4e8;border-radius:6px;overflow:hidden;">
      <thead>
        <tr style="background:#f6f8fa;">
          <th style="padding:8px 12px;text-align:left;font-size:12px;color:#586069;border-bottom:1px solid #e1e4e8;">INDICATEUR</th>
          <th style="padding:8px 12px;font-size:12px;color:#586069;border-bottom:1px solid #e1e4e8;">DERNIÈRE VALEUR</th>
          <th style="padding:8px 12px;font-size:12px;color:#586069;border-bottom:1px solid #e1e4e8;">STATUT</th>
        </tr>
      </thead>
      <tbody>{rows_all}</tbody>
    </table>
  </td></tr>

  <!-- Footer -->
  <tr><td style="background:#f6f8fa;padding:16px 28px;border-top:1px solid #e1e4e8;">
    <p style="margin:0;font-size:12px;color:#586069;text-align:center;">
      Alerte générée automatiquement par <b>CyberDashboard Alert Engine</b> — {company}<br>
      Seuils configurables dans l'onglet <i>Configuration</i> du fichier Excel.<br>
      Ne pas répondre à cet email.
    </p>
  </td></tr>

</table>
</td></tr>
</table>
</body></html>"""
    return html


# ─────────────────────────────────────────────────────────────────────────────
# Mail sending
# ─────────────────────────────────────────────────────────────────────────────

def send_alert_email(cfg: dict, alerts: list, all_data: list, dry_run: bool):
    html = build_html_email(alerts, cfg["company"], all_data)
    has_critical = any(a["level"] == "CRITIQUE" for a in alerts)
    subject_prefix = "[CRITIQUE]" if has_critical else "[AVERTISSEMENT]"
    subject = f"{subject_prefix} CyberDashboard — {len(alerts)} alerte(s) détectée(s) — {cfg['company']}"

    if dry_run:
        print("\n" + "─" * 60)
        print(f"DRY-RUN — Mail qui SERAIT envoyé :")
        print(f"  De      : {cfg['sender']}")
        print(f"  À       : {cfg['recipient']}")
        print(f"  Sujet   : {subject}")
        print(f"  Alertes : {len(alerts)}")
        for a in alerts:
            print(f"    • {a['name']} = {a['value']:,.0f} ({a['level']})")
        print("─" * 60)
        return

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = cfg["sender"]
    msg["To"]      = cfg["recipient"]
    msg["X-Priority"] = "1" if has_critical else "3"
    msg.attach(MIMEText(html, "html", "utf-8"))

    try:
        with smtplib.SMTP(cfg["smtp_server"], cfg["smtp_port"], timeout=15) as server:
            server.ehlo()
            server.starttls()
            server.login(cfg["sender"], cfg["password"])
            server.sendmail(cfg["sender"], cfg["recipient"], msg.as_string())
        print(f"✅  Mail d'alerte envoyé à {cfg['recipient']}")
    except smtplib.SMTPAuthenticationError:
        print("❌  Erreur SMTP : authentification échouée. Vérifiez l'adresse et le mot de passe.", file=sys.stderr)
        sys.exit(1)
    except smtplib.SMTPException as e:
        print(f"❌  Erreur SMTP : {e}", file=sys.stderr)
        sys.exit(1)


# ─────────────────────────────────────────────────────────────────────────────
# Logging to Excel journal
# ─────────────────────────────────────────────────────────────────────────────

def log_alerts_to_excel(path: str, alerts: list, mail_sent: bool):
    wb = openpyxl.load_workbook(path)
    ws = wb["Journal Alertes"]
    now = datetime.datetime.now()
    next_row = ws.max_row + 1

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    def border_thin():
        s = Side(style="thin", color="30363D")
        return Border(left=s, right=s, top=s, bottom=s)

    for a in alerts:
        level = a["level"]
        bg = "F8D7DA" if level == "CRITIQUE" else "FFF3CD"
        fg = "842029" if level == "CRITIQUE" else "856404"
        sev_icon = f"🔴 {level}" if level == "CRITIQUE" else f"🟡 {level}"
        row_data = [
            now, a["name"], a["value"], a["crit"] if level == "CRITIQUE" else a["warn"],
            sev_icon, "✅ Oui" if mail_sent else "🔇 Non (dry-run)",
            f"Variation: {a['variation']:+.1f}% vs valeur précédente"
        ]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(next_row, c_idx, val)
            cell.border = border_thin()
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(name="Arial", size=10, color=fg, bold=(c_idx in (1, 5)))
            cell.alignment = Alignment(horizontal="center" if c_idx != 7 else "left", vertical="center")
            if c_idx == 1:
                cell.number_format = "DD/MM/YYYY HH:MM:SS"
        next_row += 1

    wb.save(path)
    print(f"📋  {len(alerts)} entrée(s) ajoutée(s) au Journal Alertes.")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Cyber Incident Alert Engine")
    parser.add_argument("--file",    default=DEFAULT_FILE, help="Chemin vers le fichier Excel")
    parser.add_argument("--dry-run", action="store_true",  help="Analyse sans envoi de mail")
    args = parser.parse_args()

    xlsx_path = Path(args.file)
    if not xlsx_path.exists():
        print(f"❌  Fichier introuvable : {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    print(f"🔍  Lecture de {xlsx_path} …")
    wb        = load_workbook_data(str(xlsx_path))
    cfg       = read_config(wb)
    indicators = read_thresholds(wb, INDICATORS)
    data      = read_last_two_values(wb, indicators)
    alerts    = detect_anomalies(data)

    print(f"📊  {len(data)} indicateurs analysés — {len(alerts)} alerte(s) détectée(s).")

    if not alerts:
        print("✅  Aucun seuil dépassé. Système nominal.")
        return

    for a in alerts:
        level_icon = "🔴" if a["level"] == "CRITIQUE" else "🟡"
        print(f"  {level_icon}  {a['name']}: {a['value']:,.0f} (seuil: {a['crit']:,.0f}) "
              f"| variation: {a['variation']:+.1f}%")

    mail_sent = not args.dry_run
    send_alert_email(cfg, alerts, data, dry_run=args.dry_run)
    log_alerts_to_excel(str(xlsx_path), alerts, mail_sent=mail_sent)

    print("\n✅  Traitement terminé.")


if __name__ == "__main__":
    main()
