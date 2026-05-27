"""
data_feeder.py — Simulateur de données réalistes
==================================================
Génère et injecte une nouvelle ligne de données dans l'onglet "Données"
du fichier CyberDashboard_Data.xlsx.

Simule un comportement réaliste :
  - Variations naturelles selon l'heure (pic matin/soir, creux la nuit)
  - Incidents aléatoires ponctuels (DDoS, phishing spike)
  - Dérive progressive possible (montée lente d'un scan)

Usage :
    python data_feeder.py                      # injecte 1 ligne (heure courante)
    python data_feeder.py --rows 24            # injecte 24 lignes (simulation 24h)
    python data_feeder.py --rows 7 --days      # injecte 7 jours (1 ligne/jour)
    python data_feeder.py --incident ddos      # force un pic DDoS sur cette ligne
    python data_feeder.py --incident phishing  # force un pic phishing
    python data_feeder.py --dry-run            # affiche sans écrire dans l'Excel
"""

import argparse
import datetime
import math
import random
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────────────────────
DEFAULT_FILE = "CyberDashboard_Data.xlsx"

# Profils de base par heure (facteur multiplicateur, 0h→23h)
# Simule les pics d'activité matin (9h) et soir (18h), creux la nuit
HOURLY_PROFILE = [
    0.15, 0.10, 0.08, 0.08, 0.10, 0.20,   # 0h–5h  : nuit
    0.40, 0.70, 0.90, 1.00, 0.95, 0.90,   # 6h–11h : montée matin
    0.85, 0.90, 0.88, 0.85, 0.95, 1.00,   # 12h–17h: journée
    0.90, 0.75, 0.60, 0.45, 0.30, 0.20,   # 18h–23h: descente soir
]

# Valeurs de base (activité normale à pleine charge)
BASE = {
    "mails":      1000,
    "net_req":    7000,
    "login_att":  120,
    "files":      40,
    "ids":        2,
    "traf_in":    90.0,
    "traf_out":   35.0,
    "refused":    50,
    "suspect":    1,
}

# Amplitudes de bruit aléatoire (±%)
NOISE = {
    "mails":      0.20,
    "net_req":    0.18,
    "login_att":  0.35,
    "files":      0.25,
    "ids":        0.60,
    "traf_in":    0.22,
    "traf_out":   0.20,
    "refused":    0.40,
    "suspect":    1.00,
}

# Multiplicateurs d'incident
INCIDENTS = {
    "ddos": {
        "net_req":   (4.0, 6.0),
        "traf_in":   (5.0, 9.0),
        "refused":   (4.0, 8.0),
        "ids":       (6.0, 12.0),
        "login_att": (1.0, 1.5),
        "mails":     (1.0, 1.2),
        "suspect":   (2.0, 4.0),
    },
    "phishing": {
        "mails":     (3.0, 6.0),
        "login_att": (3.0, 7.0),
        "ids":       (4.0, 8.0),
        "net_req":   (1.2, 1.8),
        "refused":   (1.5, 2.5),
        "traf_in":   (1.0, 1.3),
        "suspect":   (3.0, 6.0),
    },
    "bruteforce": {
        "login_att": (6.0, 12.0),
        "refused":   (5.0, 10.0),
        "ids":       (4.0, 8.0),
        "net_req":   (1.3, 2.0),
        "mails":     (1.0, 1.1),
        "traf_in":   (1.0, 1.2),
        "suspect":   (2.0, 5.0),
    },
    "malware": {
        "suspect":   (8.0, 15.0),
        "traf_out":  (3.0, 6.0),
        "ids":       (5.0, 10.0),
        "net_req":   (1.5, 2.5),
        "login_att": (2.0, 4.0),
        "mails":     (1.0, 1.2),
        "refused":   (1.0, 1.5),
    },
}

# ─────────────────────────────────────────────────────────────────────────────

def noisy(base: float, noise_pct: float) -> float:
    """Applique un bruit gaussien ±noise_pct% autour de base."""
    sigma = base * noise_pct / 2
    return max(0.0, random.gauss(base, sigma))


def generate_row(dt: datetime.datetime, incident: str | None = None,
                 random_incident_prob: float = 0.03) -> dict:
    """
    Génère un jeu de valeurs pour l'horodatage dt.
    incident : None | 'ddos' | 'phishing' | 'bruteforce' | 'malware'
    random_incident_prob : probabilité d'un incident spontané si incident=None
    """
    hour = dt.hour
    factor = HOURLY_PROFILE[hour]

    # Valeurs de base modulées par l'heure + bruit
    vals = {
        k: noisy(BASE[k] * factor, NOISE[k])
        for k in BASE
    }

    # Incident aléatoire spontané ?
    if incident is None and random.random() < random_incident_prob:
        incident = random.choice(list(INCIDENTS.keys()))
        print(f"  ⚡ Incident spontané simulé : {incident.upper()}")

    # Application des multiplicateurs d'incident
    if incident and incident in INCIDENTS:
        mults = INCIDENTS[incident]
        for key, (lo, hi) in mults.items():
            vals[key] *= random.uniform(lo, hi)

    # Arrondi et typage
    return {
        "datetime":   dt,
        "mails":      max(0, round(vals["mails"])),
        "net_req":    max(0, round(vals["net_req"])),
        "login_att":  max(0, round(vals["login_att"])),
        "files":      max(0, round(vals["files"])),
        "ids":        max(0, round(vals["ids"])),
        "traf_in":    max(0.0, round(vals["traf_in"], 1)),
        "traf_out":   max(0.0, round(vals["traf_out"], 1)),
        "refused":    max(0, round(vals["refused"])),
        "suspect":    max(0, round(vals["suspect"])),
        "_incident":  incident,
    }


def border_thin():
    s = Side(style="thin", color="30363D")
    return Border(left=s, right=s, top=s, bottom=s)


def write_row_to_excel(ws, row_data: dict):
    """Ajoute une ligne en bas de l'onglet Données avec le bon formatage."""
    next_row = ws.max_row + 1
    values = [
        row_data["datetime"],
        row_data["mails"],
        row_data["net_req"],
        row_data["login_att"],
        row_data["files"],
        row_data["ids"],
        row_data["traf_in"],
        row_data["traf_out"],
        row_data["refused"],
        row_data["suspect"],
    ]

    bg_light  = "F0F6FC"
    bg_dark   = "0D1117"
    text_dark = "0D1117"
    text_light = "E6EDF3"

    for c_idx, val in enumerate(values, 1):
        cell = ws.cell(next_row, c_idx, val)
        cell.border = border_thin()
        cell.alignment = Alignment(horizontal="left", vertical="center")

        if c_idx == 1:
            cell.number_format = "DD/MM/YYYY HH:MM"
            cell.font = Font(name="Arial", size=10, color=text_light)
            cell.fill = PatternFill("solid", fgColor=bg_dark)
        else:
            cell.font = Font(name="Arial", size=10, color=text_dark)
            cell.fill = PatternFill("solid", fgColor=bg_light)
            cell.number_format = "#,##0" if c_idx != 7 and c_idx != 8 else "#,##0.0"

    ws.row_dimensions[next_row].height = 20


def print_row(row_data: dict):
    inc = f"  ← 🚨 {row_data['_incident'].upper()}" if row_data["_incident"] else ""
    print(
        f"  {row_data['datetime'].strftime('%d/%m/%Y %H:%M')} | "
        f"Mails: {row_data['mails']:>5} | "
        f"NetReq: {row_data['net_req']:>6} | "
        f"Login: {row_data['login_att']:>4} | "
        f"IDS: {row_data['ids']:>3} | "
        f"Suspect: {row_data['suspect']:>2}"
        f"{inc}"
    )


# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="CyberDashboard — Simulateur de données")
    parser.add_argument("--file",     default=DEFAULT_FILE, help="Chemin vers le fichier Excel")
    parser.add_argument("--rows",     type=int, default=1,  help="Nombre de lignes à injecter")
    parser.add_argument("--days",     action="store_true",  help="Mode jours (1 ligne = 1 jour)")
    parser.add_argument("--incident", choices=["ddos","phishing","bruteforce","malware"],
                        default=None, help="Forcer un type d'incident sur la dernière ligne")
    parser.add_argument("--dry-run",  action="store_true",  help="Afficher sans écrire dans l'Excel")
    args = parser.parse_args()

    xlsx_path = Path(args.file)
    if not xlsx_path.exists():
        print(f"❌  Fichier introuvable : {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    # Déterminer le point de départ temporel
    if not args.dry_run:
        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb["Données"]
        last_dt = ws.cell(ws.max_row, 1).value
    else:
        wb = None
        ws = None
        last_dt = None

    if isinstance(last_dt, datetime.datetime):
        start_dt = last_dt
    else:
        start_dt = datetime.datetime.now().replace(minute=0, second=0, microsecond=0)

    step = datetime.timedelta(days=1) if args.days else datetime.timedelta(hours=1)

    print(f"\n📥  Injection de {args.rows} ligne(s) dans {xlsx_path}")
    print(f"    Mode    : {'Jours' if args.days else 'Heures'}")
    print(f"    Incident: {args.incident.upper() if args.incident else 'aléatoire (3%)'}")
    print(f"    Dry-run : {'OUI' if args.dry_run else 'NON'}")
    print()

    generated = []
    for i in range(args.rows):
        current_dt = start_dt + step * (i + 1)
        # L'incident forcé s'applique uniquement à la dernière ligne
        inc = args.incident if (i == args.rows - 1 and args.incident) else None
        row = generate_row(current_dt, incident=inc)
        generated.append(row)
        print_row(row)

    if args.dry_run:
        print("\n🔇  Dry-run : aucune écriture effectuée.")
        return

    for row in generated:
        write_row_to_excel(ws, row)

    wb.save(str(xlsx_path))
    print(f"\n✅  {len(generated)} ligne(s) ajoutée(s) dans '{xlsx_path}'.")
    print("    Lance maintenant : python alert_engine.py --dry-run  pour tester les alertes.")


if __name__ == "__main__":
    main()
